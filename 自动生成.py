import pandas
import openpyxl as opxl
import os
import xlwings as xw

year = ""
mon = ""
day = ""
workshop_name = ""
工区名数组 = []
clolist = ["工区", "作业日期", "负责人","线名", "行别", "点内点外", "出行方式",
           "作业项目", "作业地点", "作业时段", "影响范围", "上道位置", "下道位置", "登记站", "作业区段", "职工数",
           "劳务工", "带班人", "作业人员", "机具", "配合单位", "非工务配合单位", "补充说明", "驻站防护人", "现场防护人", "远方防护人", "中间联络人", "车间盯控人"
           ]
def 点外单赋值(工区名,点外单,计划):
    s = 4
    
    print("................")
    print(shuju)
    '''for i in range(2,计划.nrows):
        if len(计划.cell(i,1).value) > 0 and 计划.cell(i,13).value == "点外":
            if 计划.cell(i,1).value == 工区名:
                s = s + 1
                项目 = []
                点外单.cell(s,2).value = 计划.cell(i,5).value         #日期
                点外单.cell(s,3).value = 计划.cell(i,10).value        #线别
                点外单.cell(s,4).value = 计划.cell(i,24).value        #站/区
                点外单.cell(s,5).value = 计划.cell(i,19).value        #时间
                点外单.cell(s,6).value = 计划.cell(i,23).value        #登记站
                点外单.cell(s,7).value = 计划.cell(i,11).value        #行别
                点外单.cell(s,8).value = 计划.cell(i,13).value        #项目
                点外单.cell(s,9).value = 计划.cell(i,18).value        #施工里程
                点外单.cell(11,6).value = 计划.cell(i,45).value        #施工里程
                项目1 = (计划.cell(i,17).value).replace("【过渡】","")
                项目 = 项目1.split("【")            #作业项目
                for xm in 项目:
                    if type(点外单.cell(s,10).value) == type(None) :
                        点外单.cell(s,10).value = ""
                    if len(点外单.cell(s,10).value) > 3 :
                        点外单.cell(s,10).value = (点外单.cell(s,10).value) +"、"
                    点外单.cell(s,10).value = 点外单.cell(s,10).value + xm.split("（")[0]
                #点外单.cell(s,11).value = 计划.cell(i,8).value #维修类型
                点外单.cell(s,11).value = "柳州工务段：" + 计划.cell(i,6).value         #作业负责人
                if "非上线" in 计划.cell(i,12).value:
                    是否上线 = "\n三、非上线作业"
                else:
                    是否上线 = "\n三、上线作业"
                点外单.cell(s,12).value = "一、驻站联络员：" + 计划.cell(i,34).value + "\n二、现场防护员：" + 计划.cell(i,35).value + 是否上线

                if "至K" in 点外单.cell(s,9).value :
                    点外单.cell(s,9).value = (点外单.cell(s,9).value).split("至K")[0] + "m至K" + (点外单.cell(s,9).value).split("至K")[1]
                    点外单.cell(s,9).value = (点外单.cell(s,9).value).split("至K")[0] + "至K" + (点外单.cell(s,9).value).split("至K")[1][0:7] + "m" + (点外单.cell(s,9).value).split("至K")[1][7:]
                    点外单.cell(s,9).value = (点外单.cell(s,9).value).replace("K","")
                    zfc =点外单.cell(s,9).value
                    点外单.cell(s,9).value = (zfc).split("线：")[0] + "线：" + ((zfc).split("线：")[1]).replace("﹢","km")
                    点外单.cell(s,9).value = (点外单.cell(s,9).value).split("（")[0]'''
    return(点外单)

print("本程序由南丹线路车间编写!")
if len(os.listdir("请将导出的日计划放入此文件夹内\\")) == 0:
    print(
        "\n未找到日计划文件请将“XXX线路车间XXXX-XX-XX日计划”文件放入“请将导出的日计划放入此文件夹内”文件夹中后再次运行本程序！\n")
for filename in os.listdir("请将导出的日计划放入此文件夹内\\"):
    print(filename)
    year = filename.split("至")[0].split("车间 ")[1].split("-")[0]  # 从文件名中分离出年
    mon = filename.split("至")[0].split("车间 ")[1].split("-")[1]  # 从文件名中分离出月
    day = filename.split("至")[0].split("车间 ")[1].split("-")[2]  # 从文件名中分离出日
    #---------------------------核对表---------------------------------------
    模板 = opxl.load_workbook("模板.xlsx")
    日计划 = pandas.read_html("请将导出的日计划放入此文件夹内\\" + filename, encoding='UTF8', header=0, index_col=0)
    workshop_name = filename.split(" ")[1]  # 从文件名中分离出车间名
    #---------------------------核对表---------------------------------------
    五合一核对表 = 模板["五合一核对表"]
    # print(日计划[0])
    五合一核对表.cell(1,1).value = workshop_name + year + "-" + mon + "-" + day + "计划核对表"
    计划 = 日计划[0]
    计划 = 计划.fillna(value='')

    #print(计划)
    #data = 计划["{worksheet}"]
    aaa = 计划.loc[:, clolist]
    rows  = 3
    for row in aaa.values:
        
        c = 1
        #print(row)
        for col in row:
            五合一核对表.cell(rows,c).value =str(col)
            c +=1
        rows += 1
    del 模板["点内"]
    del 模板["点外"]
    del 模板["点外作业单"]
    path = os.getcwd() + "\\" + mon + "月" + day + "日五合一核对表" + ".xlsx"
    模板.save(path)

    with xw.App(visible=True, add_book=False) as app:
        workbook = app.books.open(path)
        worksheet = workbook.sheets
        for i in worksheet:
            i.autofit()
        workbook.save()
        workbook.close()
    #---------------------------点外单---------------------------------------
    模板 = opxl.load_workbook("模板.xlsx")
    点外单 = 模板["点外作业单"]
    '''for i in range(2, 计划.nrows):
        if 计划.cell(i, 1).value not in 工区名数组:
            工区名数组.insert(0, 计划.cell(i, 1).value)'''
    dwlist = ["作业日期","线名", "作业区段","作业时段","登记站","行别","点内点外","作业地点","作业项目","负责人","驻站防护人", "现场防护人","工区"]
    shuju = 计划.loc[:, dwlist]
    gq = shuju.loc[:,"工区"]
    for i in gq:    #!!!!!!!!!!!!!!!!!!!!!!!!!
        if gq.cell(i, 0).value not in 工区名数组:
            工区名数组.insert(0, gq.cell(i, 0).value)
    print(工区名数组)
    '''for gqm in 工区名数组:
        表格名称 = gqm
        cp_sheet = 模板.copy_worksheet(点外单)
        cp_sheet.title = 表格名称
        cp_sheet.cell(1, 1).value = "柳州工务段南丹线路车间" + year + "年" + mon + "月" + day + "日天窗点外作业计划表"
        cp_sheet = 点外单赋值(gqm, cp_sheet, 计划)
    del 模板["点内"]
    del 模板["点外"]
    del 模板["五合一核对表"]
    del 模板["点外作业单"]'''
    #模板.save(os.getcwd() + "\\" + workshop_name + year + "年" + mon +"月" + day +"点外作业单.xlsx")













    #os.remove("请将导出的日计划放入此文件夹内\\" + filename)




'''
    cp_sheet = 模板.copy_worksheet(Sheet)
    cp_sheet.title = dicts[name][0]
    cp_sheet.cell(1, 1).value = name + mon + "月" + day + "日作业计划表"
    for row in aaa.values:

        c = 5
        for col in row:
            cp_sheet.cell(r, c).value = str(col)
            c += 1

        if "点外" in cp_sheet.cell(r, 14).value:
            作业号 += 1
            准许号 += 1
            if len(str(作业号)) == 1:
                作业 = dicts[name][1] + mon + day + "00" + str(作业号)
            elif len(str(作业号)) == 2:
                作业 = dicts[name][1] + mon + day + "0" + str(作业号)
            elif len(str(作业号)) == 3:
                作业 = dicts[name][1] + mon + day + str(作业号)
            if len(str(准许号)) == 1:
                准许 = mon + day + "00" + str(准许号)
            elif len(str(准许号)) == 2:
                准许 = mon + day + "0" + str(准许号)
            elif len(str(准许号)) == 3:
                准许 = mon + day + str(准许号)
            cp_sheet.cell(r, 2).value = 作业
            cp_sheet.cell(r, 3).value = "Z" + 准许
            cp_sheet.cell(r, 4).value = "G" + 准许
        r += 1
    del 模板["Sheet1"]
    for sheet in 模板:
        for rows in range(52, 3, -1):
            if sheet.cell(rows, 5).value is None:
                sheet.delete_rows(rows)
    path = os.getcwd() + "\\" + "柳州工务段" + mon + "月" + day + "日作业计划表" + ".xlsx"
    模板.save(path)
    with xw.App(visible=True, add_book=False) as app:
        workbook = app.books.open(path)
        worksheet = workbook.sheets
        for i in worksheet:
            i.autofit()
        workbook.save()
        workbook.close()
    os.remove("请将导出的日计划放入此文件夹内\\" + filename)'''
input("程序已运行结束，按回车键退出：")